def import_flowers_xlsx():
    import openpyxl

    dataframe = openpyxl.load_workbook("Champ.xlsx")

    dataframe1 = dataframe.active

    list_flowers = []

    for row in range(0, dataframe1.max_row):
        pos_flowers = []
        for col in dataframe1.iter_cols(0, dataframe1.max_column):
            pos_flowers.append((col[row].value))
        list_flowers.append((pos_flowers[0], pos_flowers[1]))

    return list_flowers


def save_txt(list_flowers, name):
    with open(name + ".txt", "w") as file_txt:
        for line in list_flowers:
            file_txt.write(str(int(line[0])) + "," + str(int(line[1])) + "\n")


def import_flowers_txt(name="champ"):
    flowers = []
    with open(name + ".txt", "r") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    x, y = map(float, line.split(","))
                    flowers.append((x, y))
                except:
                    pass

    return flowers


if __name__ == "__main__":
    name = "champ"
    save_txt(import_flowers_xlsx(), name)
