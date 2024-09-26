import random
import openpyxl
from beehive import Bee, Beehive

NB_BEE = 100  # number of bees in the beehive

SELECTION_RATE = 20  # How many bees we keep per generation

MUTATION_RATE = 0.2  # How many bees will mutate

MUTATION_FREQUENCY = 2  # How many flower will change in a bee mutate path


def import_flowers():
    dataframe = openpyxl.load_workbook("Champ.xlsx")

    dataframe1 = dataframe.active

    list_flowers = []

    for row in range(1, dataframe1.max_row):
        pos_flowers = []
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            pos_flowers.append(int((col[row].value)))
        list_flowers.append((pos_flowers[0], pos_flowers[1]))

    return list_flowers


FLOWERS = import_flowers()  # List of tuple that containt flowers' positions

BEEHIVE = Beehive(NB_BEE, FLOWERS)  # Beehive


def print_bees():
    i = 0
    for b in BEEHIVE.population:
        print(f"Bee {i}  {b}")


def print_beehive():
    print(BEEHIVE)


def mutate_beehive():
    for bee in BEEHIVE.population:
        if random.random() > MUTATION_RATE:
            bee.mutation(MUTATION_FREQUENCY)


if "__main__" == __name__:
    print_beehive()
    top = BEEHIVE.selection(SELECTION_RATE)
    BEEHIVE.multiplication_population(top, NB_BEE)
    mutate_beehive()
    print_bees()
    print_beehive()
