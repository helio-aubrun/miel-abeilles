import openpyxl

def import_flowers () :
    dataframe = openpyxl.load_workbook("Champ.xlsx")

    dataframe1 = dataframe.active

    list_flowers = []

    for row in range(0, dataframe1.max_row):
        pos_flowers = []
        for col in dataframe1.iter_cols(0, dataframe1.max_column):
            pos_flowers.append ( (col[row].value))
        list_flowers.append ((pos_flowers[0],pos_flowers[1]))

    return list_flowers


def save_txt () :
    list_flowers = import_flowers ()
    with open ("Champ.txt", "w") as file_txt :
        for line in list_flowers :
            file_txt.write(str (line [0]) + "," + str (line [1]) + '\n')

def load_save () :

    file_xlsx = import_flowers ()

    save_txt ()

if __name__ == "__main__" :
    load_save  ()